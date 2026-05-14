"""
Day2 - center preprocess.py

作用：
- 读取 raw_input/selected_patients.xlsx
- 匹配 raw_input/images 与 raw_input/masks
- 生成 center/central_data_pool

边界：
- 不生成 tasks.json
- 不生成 Master_Manifest.json / Receive_Registry.json
- 不生成 results.json / merged / review_results.json / final.json
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from PIL import Image

# 允许脚本从 project_root 运行：python scripts/center/preprocess.py ...
# 也允许单文件调试时把 Day1 shared 文件放在同目录或 scripts/shared 中。
try:
    from scripts.shared.config_loader import load_config
    from scripts.shared.constants import PROJECT_ID, SCHEMA_VERSION
    from scripts.shared.id_utils import build_sample_id
except Exception:  # pragma: no cover - 兼容单文件调试
    import sys

    CURRENT = Path(__file__).resolve()
    for parent in [CURRENT.parent, *CURRENT.parents]:
        shared_dir = parent / "scripts" / "shared"
        if shared_dir.exists():
            sys.path.insert(0, str(parent))
            break
    try:
        from scripts.shared.config_loader import load_config
        from scripts.shared.constants import PROJECT_ID, SCHEMA_VERSION
        from scripts.shared.id_utils import build_sample_id
    except Exception:
        # 最小兜底：避免 Day2 预处理脚本因为 shared 包路径未配置而完全不可运行。
        def load_config(path: str) -> Dict[str, Any]:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)

        PROJECT_ID = "MED_IMG_V1"
        SCHEMA_VERSION = "v1"

        def build_sample_id(check_category: str, case_id: str, image_id: str, schema_version: str = "v1") -> str:
            if schema_version != "v1":
                raise ValueError("sample_id 仅支持 v1")
            parts = [check_category.strip(), case_id.strip(), image_id.strip()]
            if any(not p for p in parts):
                raise ValueError("sample_id fields must be non-empty strings")
            if any(("/" in p or "\\" in p) for p in parts):
                raise ValueError("sample_id fields must not contain path separators")
            return f"{parts[0]}_{parts[1]}_{parts[2]}"


ALLOWED_IMAGE_EXTS = {".jpg", ".png"}
MASK_EXT = ".png"
ISO_FORMAT = "%Y-%m-%dT%H:%M:%S"


class PreprocessError(Exception):
    """预处理阶段可记录到 error_report.json 的错误。"""


def now_iso() -> str:
    return datetime.now().strftime(ISO_FORMAT)


def to_posix(path: Path | str) -> str:
    return str(PurePosixPath(str(path).replace("\\", "/")))


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def atomic_write_json(path: Path, data: Any) -> None:
    ensure_dir(path.parent)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    os.replace(tmp, path)


def clean_text(value: Any, field_name: str, allow_null: bool = False) -> Optional[str]:
    if value is None:
        if allow_null:
            return None
        raise PreprocessError(f"{field_name} 缺失")
    text = str(value).strip()
    if text == "":
        if allow_null:
            return None
        raise PreprocessError(f"{field_name} 为空")
    return text


def get_basename_without_ext(filename: Any) -> str:
    raw = clean_text(filename, "图片文件名")
    # Excel 中不允许带目录，只接受 basename
    normalized = raw.replace("\\", "/")
    if "/" in normalized:
        raise PreprocessError(f"图片文件名不得包含目录路径: {raw}")
    return Path(normalized).stem.strip()


def index_files_by_stem(directory: Path, allowed_exts: Iterable[str], kind: str) -> Dict[str, Path]:
    if not directory.exists() or not directory.is_dir():
        raise FileNotFoundError(f"{kind} 目录不存在: {directory}")

    allowed = {ext.lower() for ext in allowed_exts}
    result: Dict[str, Path] = {}
    duplicates: Dict[str, List[str]] = defaultdict(list)

    for path in directory.iterdir():
        if not path.is_file():
            continue
        ext = path.suffix.lower()
        if ext not in allowed:
            continue
        stem = path.stem.strip()
        if stem in result:
            duplicates[stem].append(path.name)
        else:
            result[stem] = path

    if duplicates:
        detail = {k: [result[k].name, *v] for k, v in duplicates.items()}
        raise ValueError(f"{kind} 目录中存在同 basename 重复文件: {detail}")

    return result


def classify_resolution(width: int, height: int, resolution_rules: Dict[str, Any]) -> str:
    short_edge = min(width, height)
    long_edge = max(width, height)

    s_rule = resolution_rules["S"]
    if short_edge <= s_rule["short_edge_max"] and long_edge <= s_rule["long_edge_max"]:
        return "S"

    m_rule = resolution_rules["M"]
    if (
        m_rule["short_edge_min"] <= short_edge <= m_rule["short_edge_max"]
        and m_rule["long_edge_min"] <= long_edge <= m_rule["long_edge_max"]
    ):
        return "M"

    l_rule = resolution_rules["L"]
    if (
        short_edge > l_rule["short_edge_min_exclusive"]
        and long_edge > l_rule["long_edge_min_exclusive"]
    ):
        return "L"

    raise PreprocessError(
        f"resolution_level_invalid: width={width}, height={height}, short_edge={short_edge}, long_edge={long_edge}"
    )


def copy_image_to_pool(src: Path, dst: Path) -> Tuple[int, int]:
    ensure_dir(dst.parent)
    with Image.open(src) as img:
        width, height = img.size
        # 不改变原始 jpg/png 语义，只验证可打开并复制。
        shutil.copy2(src, dst)
    return width, height


def standardize_mask_to_png(src: Path, dst: Path) -> None:
    ensure_dir(dst.parent)
    with Image.open(src) as mask:
        # 保留 mask 信息，不做业务重编码；统一输出 PNG。
        mask.save(dst, format="PNG")


def generate_downsample_candidates(
    image_path: Path,
    sample_id: str,
    resolution_level: str,
    central_data_pool: Path,
    downsample_config: Dict[str, Any],
) -> Dict[str, str]:
    generate_for = downsample_config.get("generate_for", {"S": [], "M": ["x2"], "L": ["x2", "x4"]})
    scales = generate_for.get(resolution_level, [])
    generated: Dict[str, str] = {}

    if not scales:
        return generated

    with Image.open(image_path) as img:
        img_rgb = img.convert("RGB")
        width, height = img_rgb.size

        for scale in scales:
            if scale == "x2":
                factor = 2
            elif scale == "x4":
                factor = 4
            else:
                raise ValueError(f"非法 downsample scale: {scale}")

            new_size = (max(1, width // factor), max(1, height // factor))
            resized = img_rgb.resize(new_size, Image.Resampling.BICUBIC)
            out_rel = PurePosixPath("downsample_candidates") / scale / f"{sample_id}.jpg"
            out_abs = central_data_pool / out_rel
            ensure_dir(out_abs.parent)
            resized.save(out_abs, format="JPEG", quality=95)
            generated[scale] = to_posix(out_rel)

    return generated


def read_excel_rows(excel_path: Path, columns: Dict[str, str]) -> List[Dict[str, Any]]:
    if not excel_path.exists() or not excel_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel 表头为空")

    header_map = {str(v).strip(): idx for idx, v in enumerate(header_row) if v is not None and str(v).strip()}
    missing = [excel_col for excel_col in columns.values() if excel_col not in header_map]
    if missing:
        raise ValueError(f"Excel 缺少必填列: {missing}")

    rows: List[Dict[str, Any]] = []
    for excel_row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {field: row[header_map[excel_col]] if header_map[excel_col] < len(row) else None for field, excel_col in columns.items()}
        item["_row_index"] = excel_row_index
        rows.append(item)

    wb.close()
    return rows


def validate_config(config: Dict[str, Any]) -> None:
    required_top = {"project_id", "source_batch", "schema_version", "config_version", "script_version", "input", "output", "excel_columns"}
    missing = required_top - set(config.keys())
    if missing:
        raise ValueError(f"preprocess_config.json 缺少字段: {sorted(missing)}")

    if config["project_id"] != PROJECT_ID:
        raise ValueError(f"project_id 必须为 {PROJECT_ID}")
    if config["schema_version"] != SCHEMA_VERSION:
        raise ValueError(f"schema_version 必须为 {SCHEMA_VERSION}")

    for key in ["excel_path", "image_dir", "mask_dir"]:
        if key not in config["input"]:
            raise ValueError(f"input.{key} 缺失")
    if "central_data_pool" not in config["output"]:
        raise ValueError("output.central_data_pool 缺失")

    for field in ["check_category", "case_id", "image_filename", "diagnosis_raw"]:
        if field not in config["excel_columns"]:
            raise ValueError(f"excel_columns.{field} 缺失")


def preprocess(config_path: str) -> None:
    config = load_config(config_path)
    validate_config(config)

    excel_path = Path(config["input"]["excel_path"])
    image_dir = Path(config["input"]["image_dir"])
    mask_dir = Path(config["input"]["mask_dir"])
    central_data_pool = Path(config["output"]["central_data_pool"])

    images_dir = central_data_pool / "images"
    masks_dir = central_data_pool / "masks"
    metadata_dir = central_data_pool / "metadata"
    downsample_dir = central_data_pool / "downsample_candidates"
    log_dir = Path("logs") / "preprocess" / config["source_batch"]

    for p in [images_dir, masks_dir, metadata_dir, downsample_dir / "x2", downsample_dir / "x4", log_dir]:
        ensure_dir(p)

    rows = read_excel_rows(excel_path, config["excel_columns"])
    image_index = index_files_by_stem(image_dir, ALLOWED_IMAGE_EXTS, "image")
    mask_index = index_files_by_stem(mask_dir, {MASK_EXT}, "mask")

    candidates: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    # 第一遍：生成字段与 sample_id，先不入池，便于处理重复 sample_id。
    for row in rows:
        row_index = row["_row_index"]
        image_filename_for_error = row.get("image_filename")
        try:
            check_category = clean_text(row.get("check_category"), "检查分类")
            case_id = clean_text(row.get("case_id"), "检查HIS号")
            image_id = get_basename_without_ext(row.get("image_filename"))
            diagnosis_raw = clean_text(row.get("diagnosis_raw"), "检查提示")
            sample_id = build_sample_id(check_category, case_id, image_id, schema_version=config["schema_version"])
            candidates.append(
                {
                    "row_index": row_index,
                    "sample_id": sample_id,
                    "case_id": case_id,
                    "check_category": check_category,
                    "image_id": image_id,
                    "diagnosis_raw": diagnosis_raw,
                }
            )
        except Exception as exc:
            errors.append(
                {
                    "row_index": row_index,
                    "image_filename": str(image_filename_for_error) if image_filename_for_error is not None else None,
                    "sample_id": None,
                    "error_code": infer_error_code(str(exc)),
                    "error_message": str(exc),
                }
            )

    sample_counter = Counter(c["sample_id"] for c in candidates)
    duplicated_sample_ids = {sid for sid, count in sample_counter.items() if count > 1}

    samples_index: List[Dict[str, Any]] = []
    cases_map: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    resolution_summary = {"S": 0, "M": 0, "L": 0}

    for item in candidates:
        sample_id = item["sample_id"]
        row_index = item["row_index"]
        image_id = item["image_id"]

        try:
            if sample_id in duplicated_sample_ids:
                raise PreprocessError(f"sample_id_duplicate: {sample_id}")

            src_image = image_index.get(image_id)
            if src_image is None:
                raise PreprocessError(f"image_missing: {image_id}")

            src_mask = mask_index.get(image_id)
            if src_mask is None:
                raise PreprocessError(f"mask_missing: {image_id}")

            image_ext = src_image.suffix.lower()
            if image_ext not in ALLOWED_IMAGE_EXTS:
                raise PreprocessError(f"invalid_image_format: {src_image.name}")

            dst_image = images_dir / f"{sample_id}{image_ext}"
            dst_mask = masks_dir / f"{sample_id}.png"

            width, height = copy_image_to_pool(src_image, dst_image)
            standardize_mask_to_png(src_mask, dst_mask)
            resolution_level = classify_resolution(width, height, config["resolution_rules"])
            resolution_summary[resolution_level] += 1

            downsample_paths = generate_downsample_candidates(
                dst_image,
                sample_id,
                resolution_level,
                central_data_pool,
                config.get("downsample", {}),
            )

            rel_image = PurePosixPath("central_data_pool/images") / f"{sample_id}{image_ext}"
            rel_mask = PurePosixPath("central_data_pool/masks") / f"{sample_id}.png"

            sample_record = {
                "sample_id": sample_id,
                "case_id": item["case_id"],
                "check_category": item["check_category"],
                "image_id": image_id,
                "image_path": str(rel_image),
                "mask_path": str(rel_mask),
                "diagnosis_raw": item["diagnosis_raw"],
                "resolution_level": resolution_level,
                "schema_version": config["schema_version"],
                "downsample": downsample_paths if downsample_paths else None,
            }
            samples_index.append(sample_record)
            cases_map[(item["case_id"], item["check_category"])].append(sample_id)

        except Exception as exc:
            errors.append(
                {
                    "row_index": row_index,
                    "image_filename": f"{image_id}",
                    "sample_id": sample_id,
                    "error_code": infer_error_code(str(exc)),
                    "error_message": str(exc),
                }
            )

    samples_index.sort(key=lambda x: x["sample_id"])

    cases_index = []
    for (case_id, check_category), sample_ids in cases_map.items():
        sorted_ids = sorted(sample_ids)
        cases_index.append(
            {
                "case_id": case_id,
                "check_category": check_category,
                "sample_ids": sorted_ids,
                "sample_count": len(sorted_ids),
                "schema_version": config["schema_version"],
            }
        )
    cases_index.sort(key=lambda x: x["case_id"])

    manifest = {
        "project_id": config["project_id"],
        "source_batch": config["source_batch"],
        "schema_version": config["schema_version"],
        "config_version": config["config_version"],
        "script_version": config["script_version"],
        "created_at": now_iso(),
        "input_excel": config["input"]["excel_path"],
        "input_image_dir": config["input"]["image_dir"],
        "input_mask_dir": config["input"]["mask_dir"],
        "output_dir": to_posix(central_data_pool),
        "total_excel_rows": len(rows),
        "valid_samples": len(samples_index),
        "invalid_samples": len(errors),
        "resolution_summary": resolution_summary,
        "downsample_policy": config.get("downsample", {}).get("generate_for", {"S": [], "M": ["x2"], "L": ["x2", "x4"]}),
    }

    atomic_write_json(metadata_dir / "samples_index.json", samples_index)
    atomic_write_json(metadata_dir / "cases_index.json", cases_index)
    atomic_write_json(metadata_dir / "preprocess_manifest.json", manifest)
    atomic_write_json(log_dir / "error_report.json", errors)

    print(f"[OK] samples_index: {metadata_dir / 'samples_index.json'}")
    print(f"[OK] valid_samples={len(samples_index)}, invalid_samples={len(errors)}")
    print(f"[OK] error_report: {log_dir / 'error_report.json'}")


def infer_error_code(message: str) -> str:
    mapping = {
        "Excel 缺少必填列": "excel_missing_required_column",
        "检查HIS号": "case_id_missing",
        "检查分类": "check_category_missing",
        "图片文件名": "image_id_missing",
        "检查提示": "diagnosis_raw_missing_for_caption",
        "image_missing": "image_missing",
        "mask_missing": "mask_missing",
        "sample_id_duplicate": "sample_id_duplicate",
        "invalid_image_format": "invalid_image_format",
        "resolution_level_invalid": "resolution_level_invalid",
    }
    for key, code in mapping.items():
        if key in message:
            return code
    return "unknown_error"


def main() -> None:
    parser = argparse.ArgumentParser(description="Day2 center preprocess.py")
    parser.add_argument("--config", default="configs/preprocess_config.json", help="preprocess_config.json 路径")
    args = parser.parse_args()
    preprocess(args.config)


if __name__ == "__main__":
    main()
